import sys
sys.path.append('/Users/Brandon/Documents/Python/custom_modules')  # noqa
from ExcelColumnHeaders import getColumnHeaders
from collections import OrderedDict
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
from openpyxl import *


def parseTextMessages(inputFileName):
    r = open(inputFileName, "r")
    # open text file with read permissions
    w = open("pasteStuff.txt", "w")
    # write or overwrite text file to paste data once it's organized

    allMessages = []
    # Create an empty list that will eventual contain
    # each line of data as a dictionary

    for line in r.readlines():
        message = OrderedDict()
        line = line.split(",")

        symbols = "{ } [] \" ' "

        for sym in symbols:
            line = str(line).replace(sym, "")

        # Replace the symbols we don't care about with empty space...

        dataPoints = line.split(",")
        # create a list containg the dataPoints of each message
        # Here is a partial example of dataPoints:

        # ['message_id:m_mid.152738483922772:bfc38823745d0360587',
        # 'sender:user_id:839479384389478', 'uid:1004934894938452',
        # 'first_name:e', 'middle_name:', 'last_name:Smith',
        # 'is_pushable:true', 'is_messenger_user:true',
        # 'is_partial:false', 'email:863478687463874@facebook.com',
        # 'contact_email:null', 'phones:null', etc.... ]
        for point in dataPoints:
            point = point.split(':', 1)
            # split each point into key, val
            # examples: [' body', 'Hello world']
            #          [' action_id', '146483847893000007']
            #          [' email', '863478687463874@facebook.com']

            try:
                message[point[0].replace(" ", "")] = point[1]
                # creates a proper OrderedDict for each message
                # Here is a partial example:

                # OrderedDict([
                # ('message_id', 'm_mid.152738483922772:bfc38823745d0360587'),
                # ('sender', 'user_id:839479384389478'),
                # ('uid', '1004934894938452'),
                # ('first_name', 'Bobby'),
                # ('middle_name', ''),
                # ('last_name', 'Smith'),
                # ('is_pushable', 'true'),
                # ('is_messenger_user', 'true'),
                # ('is_partial', 'false'),
                # ('email', '863478687463874@facebook.com'),])
            except:
                message[point[0].replace(" ", "")] = ""
                # If a particular data point is empty for this message,
                # keep it that way in the conversion to a dictionary
        allMessages.append(message)

    properties = ["message_id",
                  "sender",
                  "uid",
                  "first_name",
                  "middle_name",
                  "last_name",
                  "email",
                  "username",
                  "name",
                  "body",
                  "message_state",
                  "read_by_users_ids",
                  "timestamp",
                  "send_timestamp",
                  "attachment_map",
                  "thread_id"]
    # Above is the list of the properties we care about. We will trash the all others. # noqa

    textMessages = []
    for message in allMessages:
        slimMessage = OrderedDict()
        for prop in properties:
            slimMessage[prop] = message[prop]
            # Copies properties from list above. I.E. The ones we care about.
        textMessages.append(slimMessage)
    for message in textMessages:
        for key, val in message.items():
            w.write(key + " : " + val + "\n")
            # Write to text file just for debugging purposes.
        w.write("\n")
    return textMessages
    # returns the list of dictionaries,
    # where each dictionary represents a text message


def fontProperties(header=False):

    if(header):
        styleProps = Font(name='Calibri',
                          size=16,
                          bold=True,
                          underline='single')
    else:
        styleProps = Font(name='Calibri',
                          size=16)
    return styleProps

    # Set font properties for excel spreadsheet


def writeTextToExcel(messages, outputFileName):

    wb = Workbook()  # load workbook
    ws = wb.active  # select worksheet
    ws.title = outputFileName.replace(".xlsx", "")

    headerStyle = fontProperties(header=True)
    # set font properties for the header
    regularStyle = fontProperties()
    # set font properties for other cells

    keys = list(messages[0].keys())
    # get the list of keys from one of the messages.

    headers = {}
    # dictionary to store the excel sheet headers. Useful later,
    # so we can look up columns by name instead of number.

    for i in range(1, len(keys) + 1):
        headerCell = ws.cell(row=1, column=i, value=keys[i-1])
        # Write the keys of the Dictionary 'message' to the first row of the sheet # noqa
        headerCell.font = headerStyle  # set font

    headers = getColumnHeaders(1, outputFileName)  # store headers

    r = 2
    for message in messages:  # iterate through each dictionary
        for key, val in message.items():
            # unpack keys and vals from the dictionary
            ws.cell(row=r, column=headers[key], value=val)
            # write each value of the dictionary to the proper column,
            # based on the key/header
        r += 1

    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 45
        # set width of cells

    wb.save(filename=outputFileName)  # save the workbook

texts = parseTextMessages(inputFileName="stuff.txt")
writeTextToExcel(messages=texts, outputFileName="textMessages.xlsx")
